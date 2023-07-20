import os
from django.db import models
from django.core.exceptions import ValidationError
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from document_generator.celery import app
from django_celery_beat.models import PeriodicTask, IntervalSchedule
from django.dispatch import receiver
from django.db.models.signals import post_save


class Documento(models.Model):
    nombre = models.CharField(max_length=100)
    placa = models.CharField(max_length=10)
    entidad = models.IntegerField()
    contenido = models.TextField()
    generado = models.BooleanField(default=False)
    entregado = models.BooleanField(default=False)  
    formato = models.CharField(max_length=10, choices=(('txt', 'Texto'), ('docx', 'DOCX'), ('xlsx', 'XLSX')), default='txt')

    def __str__(self):
        return self.nombre

    def generar_contenido(self):
        contenido = "Se remite a Sr(a) {{nombre}}. la disposición de presentarse en la entidad {{entidad}} con su vehículo de placa {{placa}}."
        contenido = contenido.replace('{{nombre}}', self.nombre).replace('{{entidad}}', str(self.entidad)).replace('{{placa}}', self.placa)
        return contenido

    def generar_documento_txt(self):
        contenido = self.generar_contenido()

        # Obtener la ruta completa del archivo
        file_dir = os.path.join('documentos_generados')
        os.makedirs(file_dir, exist_ok=True)  # Crear la carpeta si no existe
        file_path = os.path.join(file_dir, f"{self.nombre}.txt")

        try:
            with open(file_path, 'w') as file:
                file.write(contenido)
            self.contenido = contenido
            self.generado = True
            self.save()
            print("Documento TXT generado correctamente.")
            print("Ruta del archivo:", file_path)
            return file_path
        except Exception as e:
            print("Error al generar el documento TXT:", str(e))
            return None

    def generar_documento_docx(self):
        document = Document()
        contenido = self.generar_contenido()
        document.add_paragraph(contenido)
        
        # Obtener la ruta completa del archivo
        file_dir = os.path.join('documentos_generados')
        os.makedirs(file_dir, exist_ok=True)  # Crear la carpeta si no existe
        file_path = os.path.join(file_dir, f"{self.nombre}.docx")

        try:
            document.save(file_path)
            self.contenido = contenido
            self.generado = True
            self.save()
            print("Documento DOCX generado correctamente.")
            print("Ruta del archivo:", file_path)
            return file_path
        except Exception as e:
            print("Error al generar el documento DOCX:", str(e))
            return None

    def generar_documento_xlsx(self):
        workbook = Workbook()
        sheet = workbook.active
        contenido = self.generar_contenido()
        sheet['A1'] = contenido

        if self.entidad:
            fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            sheet['A1'].fill = fill
        
        # Obtener la ruta completa del archivo
        file_dir = os.path.join('documentos_generados')
        os.makedirs(file_dir, exist_ok=True)  # Crear la carpeta si no existe
        file_path = os.path.join(file_dir, f"{self.nombre}.xlsx")

        try:
            workbook.save(file_path)
            self.contenido = contenido
            self.generado = True
            self.save()
            print("Documento XLSX generado correctamente.")
            print("Ruta del archivo:", file_path)
            return file_path
        except Exception as e:
            print("Error al generar el documento XLSX:", str(e))
            return None

@app.task
def generar_documento_task(documento_id):
    try:
        documento = Documento.objects.get(id=documento_id)
        
        # Verificar si el documento ya ha sido entregado
        if documento.entregado:
            print(f"Documento con ID {documento_id} ya ha sido entregado.")
            return

        if documento.formato == 'txt':
            documento.generar_documento_txt()
        elif documento.formato == 'docx':
            documento.generar_documento_docx()
        elif documento.formato == 'xlsx':
            documento.generar_documento_xlsx()
        
        # Marcar el documento como entregado
        documento.entregado = True
        documento.save()

        # Si se está usando la secuencia, marca la tarea actual como completada y programa la siguiente
        if SecuenciaTarea.objects.filter(documento=documento).exists():
            secuencia_actual = SecuenciaTarea.objects.get(documento=documento, completado=False)
            secuencia_actual.completado = True
            secuencia_actual.save()

            # Programar la siguiente tarea en la secuencia
            siguiente_orden = secuencia_actual.orden + 1
            tarea = Tarea.objects.get(documento=documento)
            tarea.programar_tarea_secuencia(orden=siguiente_orden)

    except Documento.DoesNotExist:
        print(f"Documento con ID {documento_id} no encontrado.")
    except Exception as e:
        print(f"Error al procesar el documento con ID {documento_id}: {str(e)}")


class Tarea(models.Model):
    documento = models.ForeignKey(Documento, on_delete=models.CASCADE)
    fecha_generacion = models.DateTimeField(auto_now_add=True)
    intervalo = models.IntegerField(null=True, blank=True)
    periodicidad_dias = models.IntegerField(null=True, blank=True)

    def clean(self):
        if self.intervalo and self.periodicidad_dias:
            raise ValidationError("Solo puede proporcionar 'intervalo' o 'periodicidad_dias', no ambos.")

    def __str__(self):
        return f"Tarea {self.id} - Documento {self.documento.nombre}"

    def programar_tarea(self):
        if self.intervalo:
            # Lógica para programar tarea basada en el intervalo
            schedule, created = IntervalSchedule.objects.get_or_create(
                every=self.intervalo,
                period=IntervalSchedule.DAYS,
            )
            task_name = f'Generar documento {self.documento.id}'
            task, created = PeriodicTask.objects.get_or_create(
                interval=schedule,
                name=task_name,
                defaults={
                    'task': 'api.models.generar_documento_task',
                    'args': str([self.documento.id]),
                }
            )
            if not created:
                task.args = str([self.documento.id])
                task.interval = schedule
                task.save()
        elif self.periodicidad_dias:
            # Lógica para programar tarea basada en la secuencia
            self.programar_tarea_secuencia(orden=1)


    def programar_tarea_secuencia(self, orden):
        try:
            secuencia = SecuenciaTarea.objects.get(documento=self.documento, orden=orden)
            schedule, created = IntervalSchedule.objects.get_or_create(
                every=secuencia.dias_desde_anterior,
                period=IntervalSchedule.DAYS,
            )
            task_name = f'Generar documento {self.documento.id} - Secuencia {orden}'
            task, created = PeriodicTask.objects.get_or_create(
                interval=schedule,
                name=task_name,
                defaults={
                    'task': 'api.models.generar_documento_task',
                    'args': str([self.documento.id]),
                }
            )
            if not created:
                task.args = str([self.documento.id])
                task.interval = schedule
                task.save()
        except SecuenciaTarea.DoesNotExist:
            print(f"Secuencia {orden} no encontrada para el documento {self.documento.id}")


class SecuenciaTarea(models.Model):
    documento = models.ForeignKey(Documento, on_delete=models.CASCADE)
    orden = models.IntegerField()  # Representa el orden de la tarea en la secuencia
    dias_desde_anterior = models.IntegerField()  # Días desde la tarea anterior
    completado = models.BooleanField(default=False)  # Indica si esta tarea en la secuencia ha sido completada

    class Meta:
        unique_together = ['documento', 'orden']  # Asegura que el orden de las tareas para un documento sea único


@receiver(post_save, sender=Tarea)
def programar_tarea(sender, instance, created, **kwargs):
    if created:
        instance.programar_tarea()

